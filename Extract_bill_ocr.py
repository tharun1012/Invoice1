# import cv2
# import os
# from paddleocr import PaddleOCR
# import pandas as pd
# import re
# import numpy as np
# from openpyxl import load_workbook
# from openpyxl.styles import Font, Alignment
#
#
# # === IMAGE PREPROCESSING FUNCTION ===
# def preprocess_bill_image_ocr(input_path, output_path, denoise_strength=5, apply_otsu=False):
#     img = cv2.imread(input_path)
#     if img is None:
#         print(f"❌ Error: Could not read image from {input_path}")
#         return False
#     gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
#     denoised = cv2.fastNlMeansDenoising(gray, None, h=denoise_strength, templateWindowSize=7, searchWindowSize=21)
#     if apply_otsu:
#         _, processed = cv2.threshold(denoised, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
#     else:
#         processed = denoised
#     os.makedirs(os.path.dirname(output_path), exist_ok=True)
#     cv2.imwrite(output_path, processed)
#     print(f"✅ Preprocessed image saved to {output_path}")
#     return True
#
#
# # === IMPROVED EXTRACT HEADER INFORMATION ===
# def extract_header_info(ocr_data):
#     header_info = {'name': '', 'sl_no': '', 'date': ''}
#     header_elements = ocr_data[:50]  # Increased from 35 to capture more
#
#     # First pass: Find Sl. No and Date (they're usually together on the right)
#     for i, item in enumerate(header_elements):
#         text = item['text'].strip()
#         text_lower = text.lower()
#         x_pos = item['x_center']
#         y_pos = item['y_center']
#
#         # Extract Sl. No - look for "sl" keyword first
#         if 'sl' in text_lower and 'no' in text_lower:
#             # Check next few elements for the number
#             for j in range(i + 1, min(i + 4, len(header_elements))):
#                 next_text = header_elements[j]['text'].strip()
#                 # Match 2-4 digit numbers
#                 if re.match(r'^\d{2,4}$', next_text) and header_elements[j]['x_center'] > 600:
#                     header_info['sl_no'] = next_text
#                     print(f"Found Sl.No: {next_text} at position {j}")
#                     break
#
#         # Also check for standalone numbers in the right region (after "sl no" text)
#         if not header_info['sl_no'] and re.match(r'^\d{2,4}$', text) and x_pos > 700 and y_pos < 250:
#             header_info['sl_no'] = text
#             print(f"Found Sl.No (standalone): {text}")
#
#         # Extract Date - look near "date" keyword and on the right side
#         if 'date' in text_lower:
#             for j in range(i + 1, min(i + 5, len(header_elements))):
#                 next_text = header_elements[j]['text'].strip()
#                 # More flexible date pattern - handles |, /, -, . separators and OCR errors
#                 date_match = re.search(r'(\d{1,2}[\|/\-\.\s]?\d{1,2}[\|/\-\.\s]?\d{2,4})', next_text)
#                 if date_match and header_elements[j]['x_center'] > 600:
#                     raw_date = date_match.group(1)
#                     # Clean up: replace | and spaces with /
#                     clean_date = re.sub(r'[\|\s]', '/', raw_date)
#                     header_info['date'] = clean_date
#                     print(f"Found Date: {clean_date} (raw: {raw_date})")
#                     break
#
#         # Also check for date pattern in current text (right side, upper portion)
#         if not header_info['date'] and x_pos > 650 and y_pos < 250:
#             date_match = re.search(r'(\d{1,2}[\|/\-\.\s]?\d{1,2}[\|/\-\.\s]?\d{2,4})', text)
#             if date_match:
#                 raw_date = date_match.group(1)
#                 clean_date = re.sub(r'[\|\s]', '/', raw_date)
#                 header_info['date'] = clean_date
#                 print(f"Found Date (direct): {clean_date}")
#
#     # Second pass: Find Name (left side, excluding noise)
#     for i, item in enumerate(header_elements):
#         text = item['text'].strip()
#         text_lower = text.lower()
#         x_pos = item['x_center']
#         y_pos = item['y_center']
#
#         # Look for "name" keyword
#         if 'name' in text_lower:
#             # Check if name is after colon in same text
#             if ':' in text:
#                 parts = text.split(':', 1)
#                 if len(parts) > 1:
#                     potential_name = parts[1].strip()
#                     # Validate it's not just dots/spaces/noise
#                     if (len(potential_name) > 2 and
#                             not re.match(r'^[.\-_\s…]+$', potential_name) and
#                             not re.match(r'^\d+$', potential_name)):
#                         header_info['name'] = potential_name
#                         print(f"Found Name (after colon): {potential_name}")
#                         break
#
#             # Check next elements for name
#             if not header_info['name']:
#                 for j in range(i + 1, min(i + 4, len(header_elements))):
#                     next_text = header_elements[j]['text'].strip()
#                     # Must be on left side, reasonable length, not dots/numbers
#                     if (header_elements[j]['x_center'] < 500 and
#                             len(next_text) > 2 and
#                             not re.match(r'^[.\-_\s…]+$', next_text) and
#                             not re.match(r'^\d+$', next_text)):
#                         header_info['name'] = next_text
#                         print(f"Found Name (after 'name' keyword): {next_text}")
#                         break
#
#     # Third pass: Fallback name search in upper-left region
#     if not header_info['name']:
#         for item in header_elements:
#             text = item['text'].strip()
#             text_lower = text.lower()
#             x_pos = item['x_center']
#             y_pos = item['y_center']
#
#             # Strict criteria for fallback name
#             exclude_keywords = ['sl', 'no', 'date', 'bill', 'contact', 'email', 'phone',
#                                 'www', 'glass', 'concepts', 'darpan', 'mrp', 'particulars',
#                                 'qty', 'rate', 'total']
#
#             if (150 < x_pos < 450 and  # Left-center region
#                     100 < y_pos < 300 and  # Upper portion
#                     len(text) > 2 and
#                     not any(kw in text_lower for kw in exclude_keywords) and
#                     not re.match(r'^[.\-_\s…]+$', text) and  # Not just dots/dashes
#                     not re.match(r'^\d+$', text) and  # Not just numbers
#                     not re.match(r'^[.,;:]+$', text)):  # Not just punctuation
#
#                 # Prefer longer, more substantive text
#                 if not header_info['name'] or len(text) > len(header_info['name']):
#                     header_info['name'] = text
#                     print(f"Found Name (fallback): {text} at x={x_pos:.1f}, y={y_pos:.1f}")
#
#     return header_info
#
#
# # === CONFIG ===
# input_image = 'C:/Users/lenovo/Downloads/bill1.jpeg'
# preprocessed_image = 'C:/Users/lenovo/Downloads/bill1_ocr_ready2.jpeg'
# save_dir = 'C:/Users/lenovo/Desktop/bill1-output'
# os.makedirs(save_dir, exist_ok=True)
#
# # === STEP 1: PREPROCESS IMAGE ===
# print("=== Starting Image Preprocessing ===")
# preprocess_success = preprocess_bill_image_ocr(input_image, preprocessed_image, denoise_strength=5, apply_otsu=False)
#
# if not preprocess_success:
#     print("❌ Preprocessing failed! Exiting.")
#     exit(1)
#
# # === STEP 2: INIT OCR ===
# print("\n=== Initializing PaddleOCR ===")
# ocr = PaddleOCR(use_angle_cls=True, lang='en')
#
# # === STEP 3: RUN OCR ===
# print("\n=== Running OCR ===")
# results = ocr.predict(preprocessed_image)
#
# # === Extract text with coordinates ===
# print("=== Extracting OCR Data ===")
# ocr_data = []
#
# if isinstance(results, list) and len(results) > 0:
#     ocr_result = results[0]
#     if isinstance(ocr_result, dict) and 'rec_texts' in ocr_result and 'dt_polys' in ocr_result:
#         texts = ocr_result['rec_texts']
#         polys = ocr_result['dt_polys']
#         scores = ocr_result.get('rec_scores', [1.0] * len(texts))
#
#         for i, (text, poly) in enumerate(zip(texts, polys)):
#             confidence = scores[i] if i < len(scores) else 1.0
#             x_coords = [point[0] for point in poly]
#             y_coords = [point[1] for point in poly]
#             x_min, x_max = min(x_coords), max(x_coords)
#             y_min, y_max = min(y_coords), max(y_coords)
#
#             ocr_data.append({
#                 'text': text.strip(),
#                 'confidence': confidence,
#                 'x_min': x_min,
#                 'x_max': x_max,
#                 'y_min': y_min,
#                 'y_max': y_max,
#                 'x_center': (x_min + x_max) / 2,
#                 'y_center': (y_min + y_max) / 2,
#                 'width': x_max - x_min,
#                 'height': y_max - y_min
#             })
#
# if not ocr_data:
#     print("❌ No OCR data extracted!")
#     exit(1)
#
# ocr_data.sort(key=lambda x: (x['y_center'], x['x_center']))
# print(f"Total OCR elements: {len(ocr_data)}")
#
# # === DEBUG: Print elements in header region ===
# print("\n=== Elements in header region (y < 300) ===")
# for i, item in enumerate(ocr_data[:50]):
#     if item['y_center'] < 300:
#         print(f"{i:2d}: x={item['x_center']:6.1f} y={item['y_center']:6.1f} | '{item['text']}'")
#
# # === EXTRACT HEADER INFORMATION ===
# print("\n=== Extracting Header Information ===")
# header_info = extract_header_info(ocr_data)
# print(f"\n✅ Extracted Header:")
# print(f"Name: '{header_info['name']}'")
# print(f"Sl. No: '{header_info['sl_no']}'")
# print(f"Date: '{header_info['date']}'")
#
# print("\n=== First 30 OCR Elements ===")
# for i, item in enumerate(ocr_data[:30]):
#     print(f"{i:2d}: x={item['x_center']:6.1f} y={item['y_center']:6.1f} | '{item['text']}'")
#
#
# # === FIND TABLE HEADER ===
# def find_table_start(data):
#     for i, item in enumerate(data):
#         text = item['text'].lower().strip()
#         if 'particulars' in text or ('qty' in text and 'rate' in text):
#             print(f"\nFound table header at element {i}: '{item['text']}'")
#             return i + 5
#     return 15
#
#
# table_start = find_table_start(ocr_data)
# table_data = ocr_data[table_start:]
# print(f"\nTable data starts at element {table_start}, {len(table_data)} elements remaining")
#
#
# # === GROUP INTO ROWS ===
# def group_into_rows(data, y_threshold=25):
#     if not data:
#         return []
#     data_sorted = sorted(data, key=lambda x: x['y_center'])
#     rows = []
#     current_row = [data_sorted[0]]
#     last_y = data_sorted[0]['y_center']
#
#     for item in data_sorted[1:]:
#         if abs(item['y_center'] - last_y) <= y_threshold:
#             current_row.append(item)
#         else:
#             if current_row:
#                 current_row.sort(key=lambda x: x['x_center'])
#                 rows.append(current_row)
#             current_row = [item]
#             last_y = item['y_center']
#
#     if current_row:
#         current_row.sort(key=lambda x: x['x_center'])
#         rows.append(current_row)
#     return rows
#
#
# table_rows = group_into_rows(table_data)
# print(f"\nGrouped into {len(table_rows)} rows")
#
#
# # === SPLIT QTY AND RATE IF COMBINED ===
# def split_qty_rate(text):
#     if not text or text.strip() == '':
#         return '', ''
#     text = text.strip()
#     if '  ' in text:
#         parts = re.split(r'\s{2,}', text)
#         if len(parts) >= 2:
#             return parts[0].strip(), ' '.join(parts[1:]).strip()
#     match = re.match(r'^(\d+[a-zA-Z]*)[€$£¥](\d+)$', text)
#     if match:
#         return match.group(1), match.group(2)
#     match = re.match(r'^(\d+[a-zA-Z]+)(\d+)$', text)
#     if match:
#         return match.group(1), match.group(2)
#     if ' ' in text:
#         parts = text.split()
#         if len(parts) >= 2:
#             return parts[0], ' '.join(parts[1:])
#     if re.match(r'^\d+[a-zA-Z]+$', text):
#         return text, ''
#     return text, ''
#
#
# # === ASSIGN COLUMNS BASED ON X-POSITION ===
# def assign_to_columns(row_elements):
#     has_typical_particulars = any(150 <= elem['x_center'] < 500 for elem in row_elements)
#     columns = {'mrp': '', 'particulars': '', 'qty_rate': '', 'total': ''}
#     items_500_660 = []
#     items_660_850 = []
#
#     for elem in row_elements:
#         x = elem['x_center']
#         text = elem['text'].strip()
#         if x < 150:
#             columns['mrp'] = columns['mrp'] + ' ' + text if columns['mrp'] else text
#         elif x < 500:
#             columns['particulars'] = columns['particulars'] + ' ' + text if columns['particulars'] else text
#         elif x < 660:
#             items_500_660.append(text)
#         elif x < 850:
#             items_660_850.append(text)
#         else:
#             columns['total'] = columns['total'] + ' ' + text if columns['total'] else text
#
#     if not has_typical_particulars and items_500_660:
#         columns['particulars'] = items_500_660[0]
#         if len(items_500_660) > 1:
#             columns['qty_rate'] = ' '.join(items_500_660[1:])
#     else:
#         if items_500_660:
#             columns['qty_rate'] = ' '.join(items_500_660)
#
#     if items_660_850:
#         if columns['qty_rate']:
#             columns['qty_rate'] = columns['qty_rate'] + ' ' + ' '.join(items_660_850)
#         else:
#             columns['qty_rate'] = ' '.join(items_660_850)
#
#     columns = {k: v.strip() for k, v in columns.items()}
#     qty, rate = split_qty_rate(columns['qty_rate'])
#     return [columns['mrp'], columns['particulars'], qty, rate, columns['total']]
#
#
# # === PROCESS ROWS INTO TABLE ===
# column_names = ['MRP', 'Particulars', 'Qty', 'Rate', 'Total']
# processed_rows = []
#
# print(f"\n=== Processing {len(table_rows)} Rows ===")
#
# for row_idx, row_elements in enumerate(table_rows):
#     row_text = ' '.join([elem['text'] for elem in row_elements]).lower()
#     if any(header in row_text for header in ['particulars', 'qty', 'rate', 'total']) and row_idx < 3:
#         print(f"Row {row_idx}: SKIPPED (Header) - {row_text[:50]}")
#         continue
#     if any(footer in row_text for footer in ['signature', 'total']) and 'sub' not in row_text:
#         if row_text.count('total') > 0 and row_text.count('sub') == 0:
#             print(f"Row {row_idx}: SKIPPED (Footer) - {row_text[:50]}")
#             continue
#     if len(row_text.strip()) < 2:
#         continue
#
#     row_data = assign_to_columns(row_elements)
#     if row_data[1] or row_data[4]:
#         processed_rows.append(row_data)
#         print(f"Row {len(processed_rows):2d}: Part='{row_data[1][:25]:25}' | Qty='{row_data[2][:8]:8}'")
#
# # === CREATE AND SAVE DATAFRAME ===
# if processed_rows:
#     df = pd.DataFrame(processed_rows, columns=column_names)
#     df = df.replace('', np.nan)
#     df_to_save = df[['Particulars', 'Qty']].copy()
#
#     excel_path = os.path.join(save_dir, 'extracted_bill_1.xlsx')
#     csv_path = os.path.join(save_dir, 'extracted_bill.csv')
#
#     # Create Excel with header row
#     with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
#         header_row = pd.DataFrame([[
#             f"Name: {header_info['name']}", '',
#             f"Sl. No: {header_info['sl_no']}", '',
#             f"Date: {header_info['date']}"
#         ]], columns=['A', 'B', 'C', 'D', 'E'])
#         header_row.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=0)
#         pd.DataFrame([['', '', '', '', '']], columns=['A', 'B', 'C', 'D', 'E']).to_excel(
#             writer, sheet_name='Sheet1', index=False, header=False, startrow=1)
#         df_to_save.to_excel(writer, sheet_name='Sheet1', index=False, startrow=2)
#
#     # Format Excel
#     wb = load_workbook(excel_path)
#     ws = wb.active
#     for col in range(1, 6):
#         cell = ws.cell(row=1, column=col)
#         cell.font = Font(bold=True)
#         cell.alignment = Alignment(horizontal='left')
#     for col in range(1, 3):
#         ws.cell(row=3, column=col).font = Font(bold=True)
#     wb.save(excel_path)
#
#     # Save CSV
#     with open(csv_path, 'w', encoding='utf-8') as f:
#         f.write(f"Name: {header_info['name']},,,Sl. No: {header_info['sl_no']},,,Date: {header_info['date']}\n\n")
#         df_to_save.to_csv(f, index=False)
#
#     print(f"\n✅ EXTRACTION COMPLETE!")
#     print(f"→ Excel: {excel_path}")
#     print(f"→ CSV: {csv_path}")
#     print(f"\n=== Header Information ===")
#     print(f"Name: {header_info['name']}")
#     print(f"Sl. No: {header_info['sl_no']}")
#     print(f"Date: {header_info['date']}")
#
#     print(f"\n=== Data Quality Analysis ===")
#     for col in ['Particulars', 'Qty']:
#         non_empty = df_to_save[col].notna().sum()
#         print(f"{col:15}: {non_empty:2d}/{len(df_to_save):2d} ({non_empty / len(df_to_save) * 100:5.1f}%) filled")
#
# else:
#     print("❌ No data extracted!")
#
# # === DEBUG OUTPUT ===
# debug_path = os.path.join(save_dir, 'debug_output.txt')
# with open(debug_path, 'w', encoding='utf-8') as f:
#     f.write("=== BILL OCR DEBUG REPORT ===\n\n")
#     f.write(f"Name: {header_info['name']}\n")
#     f.write(f"Sl. No: {header_info['sl_no']}\n")
#     f.write(f"Date: {header_info['date']}\n\n")
#     f.write("=== RAW OCR DATA (First 50 elements) ===\n")
#     for i, item in enumerate(ocr_data[:50]):
#         f.write(f"{i:3d}: x={item['x_center']:6.1f} y={item['y_center']:6.1f} | '{item['text']}'\n")
#
# print(f"\n→ Debug file: {debug_path}")

# solution:
#
import cv2
import os
from paddleocr import PaddleOCR
import pandas as pd
import re
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


# === IMAGE PREPROCESSING FUNCTION (from text1.py) ===
def preprocess_bill_image_ocr(input_path, output_path, denoise_strength=5, apply_otsu=False):
    """
    OCR-friendly preprocessing for bill images.
    Args:
        input_path: Path to input image
        output_path: Path to save processed image
        denoise_strength: Strength for denoising (default: 5, keep low)
        apply_otsu: Apply simple Otsu thresholding (True/False)
    Returns:
        True if processing succeeded, False otherwise
    """
    img = cv2.imread(input_path)
    if img is None:
        print(f"❌ Error: Could not read image from {input_path}")
        return False

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    denoised = cv2.fastNlMeansDenoising(gray, None, h=denoise_strength, templateWindowSize=7, searchWindowSize=21)

    if apply_otsu:
        _, processed = cv2.threshold(denoised, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    else:
        processed = denoised

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    cv2.imwrite(output_path, processed)
    print(f"✅ Preprocessed image saved to {output_path}")
    return True


# === EXTRACT HEADER INFORMATION (from text2.py - ROBUST VERSION) ===
def extract_header_info(ocr_data):
    """
    Robust position-based extraction of Name, Sl. No, and Date.
    Handles cases where printed labels are not detected.
    """
    header_info = {"name": "", "sl_no": "", "date": ""}

    # Limit search to top 300px
    header_zone = [item for item in ocr_data if item["y_center"] < 300]

    print("\n=== DEBUG: Header Zone Elements (y < 300) ===")
    for i, item in enumerate(header_zone):
        print(f"{i:2d}: x={item['x_center']:6.1f} y={item['y_center']:6.1f} | '{item['text']}'")

    # -----------------------------
    # 1️⃣ EXTRACT NAME (Left side)
    # -----------------------------
    name_candidates = []

    for item in header_zone:
        x, y = item['x_center'], item['y_center']
        text = item['text'].strip()
        text_lower = text.lower()

        # LEFT side: x < 300, Upper region: 80 < y < 220
        if x < 300 and 80 < y < 220:
            if not re.search(r'[a-zA-Z]', text):
                continue

            if len(text) <= 1:
                continue

            clean_text = text.replace('.', '').strip()

            exclude = ['darpan', 'glass', 'ply', 'concepts', 'email',
                       'phone', 'contact', 'www', '.com', 'sl', 'no',
                       'date', 'bill', 'mrp', 'particulars', 'qty',
                       'rate', 'total', '080', '297']

            is_noise = any(kw in text_lower for kw in exclude)

            if not is_noise and len(clean_text) >= 3:
                score = len(clean_text)
                if 40 <= x <= 150:
                    score += 5
                if 90 <= y <= 180:
                    score += 3

                name_candidates.append({
                    'text': clean_text,
                    'original': text,
                    'score': score,
                    'x': x,
                    'y': y
                })
                print(f"   👤 Name candidate: '{clean_text}' (score={score}, x={x:.1f}, y={y:.1f})")

    if name_candidates:
        best = max(name_candidates, key=lambda c: c['score'])
        header_info['name'] = best['text']
        print(f"\n✅ Selected name: '{best['text']}'")
    else:
        print("\n⚠️ No name found in left region")

    # -----------------------------
    # 2️⃣ EXTRACT SL. NO
    # -----------------------------
    for i, item in enumerate(header_zone):
        text = item['text'].strip()
        text_lower = text.lower().replace(' ', '').replace('.', '')

        if ('sl' in text_lower or 'si' in text_lower) and 'no' in text_lower:
            print(f"\n🔍 Found Sl. No. keyword: '{text}' at x={item['x_center']:.1f}")

            for j in range(i + 1, min(i + 6, len(header_zone))):
                next_item = header_zone[j]
                next_text = next_item['text'].strip()

                if re.match(r'^\d{2,6}$', next_text) and next_item['x_center'] > 700:
                    header_info['sl_no'] = next_text
                    print(f"✅ Found Sl. No: {next_text}")
                    break

            if header_info['sl_no']:
                break

    # Fallback
    if not header_info['sl_no']:
        for item in header_zone:
            if item['x_center'] > 800 and item['y_center'] < 150:
                text = item['text'].strip()
                if re.match(r'^\d{2,6}$', text):
                    header_info['sl_no'] = text
                    print(f"✅ Sl. No (fallback): {text}")
                    break

    # -----------------------------
    # 3️⃣ EXTRACT DATE
    # -----------------------------
    for item in header_zone:
        x = item['x_center']
        text = item['text'].strip()
        text_lower = text.lower()

        if 'date' in text_lower and x > 600:
            print(f"\n🔍 Found 'date' keyword: '{text}'")

            date_match = re.search(r'\.?(\d{1,2})[\|/\.\s]*(\d{1,2})[\|/\.\s]*(\d{2,4})', text)

            if date_match:
                day = date_match.group(1)
                month = date_match.group(2)
                year = date_match.group(3)

                if len(year) == 2:
                    year = '20' + year if int(year) < 50 else '19' + year

                clean_date = f"{day}/{month}/{year}"
                header_info['date'] = clean_date
                print(f"✅ Extracted date: {clean_date} (from: '{text}')")
                break

        # Fallback: Look for date-like patterns on right side
        if not header_info['date'] and x > 700 and item['y_center'] < 200:
            date_match = re.search(r'\.?(\d{1,2})[\|/\.\s]*(\d{1,2})[\|/\.\s]*(\d{2,4})', text)
            if date_match:
                day, month, year = date_match.groups()
                if len(year) == 2:
                    year = '20' + year
                clean_date = f"{day}/{month}/{year}"
                header_info['date'] = clean_date
                print(f"✅ Date (fallback): {clean_date}")
                break

    # Final fallback for date
    if not header_info['date']:
        for item in header_zone:
            text = item['text']
            if 'date' in text.lower():
                match = re.search(r'(\d{1,2})\D+(\d{1,2})\D+(\d{2,4})', text)
                if match:
                    d, m, y = match.groups()
                    if len(y) == 3:
                        y = y[-2:]
                    if len(y) == 2:
                        y = '20' + y
                    header_info['date'] = f"{d}/{m}/{y}"
                    print(f"✅ Date (second fallback): {header_info['date']} from '{text}'")
                    break

    print(f"\n{'=' * 60}")
    print(f"FINAL EXTRACTED HEADER:")
    print(f"  Name:   '{header_info['name']}'")
    print(f"  Sl. No: '{header_info['sl_no']}'")
    print(f"  Date:   '{header_info['date']}'")
    print(f"{'=' * 60}\n")

    return header_info


# === FIND TABLE HEADER (from text1.py) ===
def find_table_start(data):
    """Find where the table data starts"""
    for i, item in enumerate(data):
        text = item['text'].lower().strip()
        if 'particulars' in text or ('qty' in text and 'rate' in text):
            print(f"\nFound table header at element {i}: '{item['text']}'")
            return i + 5
    return 15


# === GROUP INTO ROWS (from text1.py) ===
def group_into_rows(data, y_threshold=25):
    """Group OCR elements into rows based on y-coordinate proximity"""
    if not data:
        return []

    data_sorted = sorted(data, key=lambda x: x['y_center'])
    rows = []
    current_row = [data_sorted[0]]
    last_y = data_sorted[0]['y_center']

    for item in data_sorted[1:]:
        if abs(item['y_center'] - last_y) <= y_threshold:
            current_row.append(item)
        else:
            if current_row:
                current_row.sort(key=lambda x: x['x_center'])
                rows.append(current_row)
            current_row = [item]
            last_y = item['y_center']

    if current_row:
        current_row.sort(key=lambda x: x['x_center'])
        rows.append(current_row)

    return rows


# === SPLIT QTY AND RATE IF COMBINED (from text1.py) ===
def split_qty_rate(text):
    """Split combined qty and rate strings"""
    if not text or text.strip() == '':
        return '', ''

    text = text.strip()

    if '  ' in text:
        parts = re.split(r'\s{2,}', text)
        if len(parts) >= 2:
            return parts[0].strip(), ' '.join(parts[1:]).strip()

    match = re.match(r'^(\d+[a-zA-Z]*)[€$£¥](\d+)$', text)
    if match:
        return match.group(1), match.group(2)

    match = re.match(r'^(\d+[a-zA-Z]+)(\d+)$', text)
    if match:
        return match.group(1), match.group(2)

    if ' ' in text:
        parts = text.split()
        if len(parts) >= 2:
            return parts[0], ' '.join(parts[1:])

    if re.match(r'^\d+[a-zA-Z]+$', text):
        return text, ''

    return text, ''


# === ASSIGN COLUMNS BASED ON X-POSITION (from text1.py) ===
def assign_to_columns(row_elements):
    """Assign elements to columns based on x-position with smart particulars detection"""
    has_typical_particulars = any(150 <= elem['x_center'] < 500 for elem in row_elements)

    columns = {
        'mrp': '',
        'particulars': '',
        'qty_rate': '',
        'total': ''
    }

    items_500_660 = []
    items_660_850 = []

    for elem in row_elements:
        x = elem['x_center']
        text = elem['text'].strip()

        if x < 150:
            columns['mrp'] = columns['mrp'] + ' ' + text if columns['mrp'] else text
        elif x < 500:
            columns['particulars'] = columns['particulars'] + ' ' + text if columns['particulars'] else text
        elif x < 660:
            items_500_660.append(text)
        elif x < 850:
            items_660_850.append(text)
        else:
            columns['total'] = columns['total'] + ' ' + text if columns['total'] else text

    if not has_typical_particulars and items_500_660:
        columns['particulars'] = items_500_660[0]
        if len(items_500_660) > 1:
            columns['qty_rate'] = ' '.join(items_500_660[1:])
    else:
        if items_500_660:
            columns['qty_rate'] = ' '.join(items_500_660)

    if items_660_850:
        if columns['qty_rate']:
            columns['qty_rate'] = columns['qty_rate'] + ' ' + ' '.join(items_660_850)
        else:
            columns['qty_rate'] = ' '.join(items_660_850)

    columns = {k: v.strip() for k, v in columns.items()}
    qty, rate = split_qty_rate(columns['qty_rate'])

    return [
        columns['mrp'],
        columns['particulars'],
        qty,
        rate,
        columns['total']
    ]


# === CONFIG ===
input_image = ''
preprocessed_image = ''
save_dir = ''
os.makedirs(save_dir, exist_ok=True)

# === STEP 1: PREPROCESS IMAGE ===
print("=== Starting Image Preprocessing ===")
preprocess_success = preprocess_bill_image_ocr(input_image, preprocessed_image, denoise_strength=5, apply_otsu=False)

if not preprocess_success:
    print("❌ Preprocessing failed! Exiting.")
    exit(1)

# === STEP 2: INIT OCR ===
print("\n=== Initializing PaddleOCR ===")
ocr = PaddleOCR(
    use_angle_cls=True,
    lang='en',
)

# === STEP 3: RUN OCR ===
print("\n=== Running OCR ===")
results = ocr.predict(preprocessed_image)

# === Extract text with coordinates ===
print("=== Extracting OCR Data ===")
ocr_data = []

if isinstance(results, list) and len(results) > 0:
    ocr_result = results[0]

    if isinstance(ocr_result, dict) and 'rec_texts' in ocr_result and 'dt_polys' in ocr_result:
        texts = ocr_result['rec_texts']
        polys = ocr_result['dt_polys']
        scores = ocr_result.get('rec_scores', [1.0] * len(texts))

        for i, (text, poly) in enumerate(zip(texts, polys)):
            confidence = scores[i] if i < len(scores) else 1.0

            x_coords = [point[0] for point in poly]
            y_coords = [point[1] for point in poly]

            x_min, x_max = min(x_coords), max(x_coords)
            y_min, y_max = min(y_coords), max(y_coords)

            ocr_data.append({
                'text': text.strip(),
                'confidence': confidence,
                'x_min': x_min,
                'x_max': x_max,
                'y_min': y_min,
                'y_max': y_max,
                'x_center': (x_min + x_max) / 2,
                'y_center': (y_min + y_max) / 2,
                'width': x_max - x_min,
                'height': y_max - y_min
            })

if not ocr_data:
    print("❌ No OCR data extracted!")
    exit(1)

ocr_data.sort(key=lambda x: (x['y_center'], x['x_center']))
print(f"Total OCR elements: {len(ocr_data)}")

# === EXTRACT HEADER INFORMATION (using text2.py logic) ===
print("\n=== Extracting Header Information ===")
header_info = extract_header_info(ocr_data)

# === PROCESS TABLE DATA (using text1.py logic) ===
print("\n=== Processing Table Data ===")
table_start = find_table_start(ocr_data)
table_data = ocr_data[table_start:]
print(f"Table data starts at element {table_start}, {len(table_data)} elements remaining")

table_rows = group_into_rows(table_data)
print(f"Grouped into {len(table_rows)} rows")

# === PROCESS ROWS INTO TABLE ===
column_names = ['MRP', 'Particulars', 'Qty', 'Rate', 'Total']
processed_rows = []

print(f"\n=== Processing {len(table_rows)} Rows ===")

for row_idx, row_elements in enumerate(table_rows):
    row_text = ' '.join([elem['text'] for elem in row_elements]).lower()

    if any(header in row_text for header in ['particulars', 'qty', 'rate', 'total']) and row_idx < 3:
        print(f"Row {row_idx}: SKIPPED (Header) - {row_text[:50]}")
        continue

    if any(footer in row_text for footer in ['signature', 'total']) and 'sub' not in row_text:
        if row_text.count('total') > 0 and row_text.count('sub') == 0:
            print(f"Row {row_idx}: SKIPPED (Footer) - {row_text[:50]}")
            continue

    if len(row_text.strip()) < 2:
        continue

    row_data = assign_to_columns(row_elements)

    if row_data[1] or row_data[4]:
        processed_rows.append(row_data)
        print(f"Row {len(processed_rows):2d}: Part='{row_data[1][:30]:30}' | Qty='{row_data[2][:10]:10}'")

# === CREATE AND SAVE DATAFRAME ===
if processed_rows:
    df = pd.DataFrame(processed_rows, columns=column_names)
    df = df.replace('', np.nan)
    df_to_save = df[['Particulars', 'Qty']].copy()

    excel_path = os.path.join(save_dir, 'extracted_bill_combined.xlsx')
    csv_path = os.path.join(save_dir, 'extracted_bill_combined.csv')

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        header_row = pd.DataFrame([[
            f"Name : {header_info['name']}",
            '',
            f"Sl. No: {header_info['sl_no']}",
            '',
            f"Date : {header_info['date']}"
        ]], columns=['A', 'B', 'C', 'D', 'E'])

        header_row.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=0)
        pd.DataFrame([['', '', '', '', '']], columns=['A', 'B', 'C', 'D', 'E']).to_excel(
            writer, sheet_name='Sheet1', index=False, header=False, startrow=1)
        df_to_save.to_excel(writer, sheet_name='Sheet1', index=False, startrow=2)

    wb = load_workbook(excel_path)
    ws = wb.active

    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left')

    for col in range(1, 3):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True)

    wb.save(excel_path)

    with open(csv_path, 'w', encoding='utf-8') as f:
        f.write(f"Name: {header_info['name']},,,Sl. No: {header_info['sl_no']},,,Date: {header_info['date']}\n")
        f.write("\n")
        df_to_save.to_csv(f, index=False)

    print(f"\n✅ EXTRACTION COMPLETE!")
    print(f"→ Total rows: {len(df_to_save)}")
    print(f"→ Excel: {excel_path}")
    print(f"→ CSV: {csv_path}")
    print(f"\n=== Header Information ===")
    print(f"Name: {header_info['name']}")
    print(f"Sl. No: {header_info['sl_no']}")
    print(f"Date: {header_info['date']}")

    print(f"\n=== Data Quality Analysis ===")
    for col in ['Particulars', 'Qty']:
        non_empty = df_to_save[col].notna().sum()
        print(f"{col:15}: {non_empty:2d}/{len(df_to_save):2d} ({non_empty / len(df_to_save) * 100:5.1f}%) filled")

    print(f"\n=== Preview ===")
    print(df_to_save.to_string(index=False))

else:
    print("❌ No data extracted!")

# #trocr+paddleocr
# import cv2
# import os
# from paddleocr import PaddleOCR
# import pandas as pd
# import re
# import numpy as np
# from openpyxl import load_workbook
# from openpyxl.styles import Font, Alignment
# from transformers import TrOCRProcessor, VisionEncoderDecoderModel
# from PIL import Image
# import torch
#
#
# # === IMAGE PREPROCESSING FUNCTION ===
# def preprocess_bill_image_ocr(input_path, output_path, denoise_strength=5, apply_otsu=False):
#     """OCR-friendly preprocessing for bill images."""
#     img = cv2.imread(input_path)
#     if img is None:
#         print(f"❌ Error: Could not read image from {input_path}")
#         return False
#
#     gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
#     denoised = cv2.fastNlMeansDenoising(gray, None, h=denoise_strength, templateWindowSize=7, searchWindowSize=21)
#
#     if apply_otsu:
#         _, processed = cv2.threshold(denoised, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
#     else:
#         processed = denoised
#
#     os.makedirs(os.path.dirname(output_path), exist_ok=True)
#     cv2.imwrite(output_path, processed)
#     print(f"✅ Preprocessed image saved to {output_path}")
#     return True
#
#
# # === TROCR INTEGRATION ===
# class TrOCRExtractor:
#     """TrOCR model for enhanced text extraction"""
#
#     def __init__(self, model_name='microsoft/trocr-base-printed'):
#         print(f"\n=== Initializing TrOCR ({model_name}) ===")
#         self.processor = TrOCRProcessor.from_pretrained(model_name)
#         self.model = VisionEncoderDecoderModel.from_pretrained(model_name)
#         self.device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
#         self.model.to(self.device)
#         print(f"✅ TrOCR loaded on {self.device}")
#
#     def extract_text_from_region(self, image, bbox):
#         """Extract text from a specific region using TrOCR"""
#         x_min, y_min, x_max, y_max = map(int, bbox)
#
#         # Crop region with padding
#         padding = 5
#         h, w = image.shape[:2] if len(image.shape) == 2 else image.shape[:2]
#         x_min = max(0, x_min - padding)
#         y_min = max(0, y_min - padding)
#         x_max = min(w, x_max + padding)
#         y_max = min(h, y_max + padding)
#
#         cropped = image[y_min:y_max, x_min:x_max]
#
#         if cropped.size == 0:
#             return ""
#
#         # Convert to PIL Image
#         if len(cropped.shape) == 2:
#             pil_image = Image.fromarray(cropped).convert('RGB')
#         else:
#             pil_image = Image.fromarray(cv2.cvtColor(cropped, cv2.COLOR_BGR2RGB))
#
#         # Process with TrOCR
#         pixel_values = self.processor(pil_image, return_tensors="pt").pixel_values.to(self.device)
#
#         with torch.no_grad():
#             generated_ids = self.model.generate(pixel_values)
#
#         text = self.processor.batch_decode(generated_ids, skip_special_tokens=True)[0]
#         return text.strip()
#
#
# # === ENSEMBLE OCR: COMBINE PADDLEOCR + TROCR ===
# def run_ensemble_ocr(image_path, preprocessed_image, use_trocr=True):
#     """
#     Run both PaddleOCR and TrOCR, then merge results for better accuracy
#     """
#     print("\n=== Running Ensemble OCR (PaddleOCR + TrOCR) ===")
#
#     # Initialize PaddleOCR
#     paddle_ocr = PaddleOCR(use_angle_cls=True, lang='en')
#
#     # Run PaddleOCR
#     print("→ Running PaddleOCR...")
#     paddle_results = paddle_ocr.predict(preprocessed_image)
#
#     # Extract PaddleOCR data
#     ocr_data = []
#     if isinstance(paddle_results, list) and len(paddle_results) > 0:
#         ocr_result = paddle_results[0]
#
#         if isinstance(ocr_result, dict) and 'rec_texts' in ocr_result:
#             texts = ocr_result['rec_texts']
#             polys = ocr_result['dt_polys']
#             scores = ocr_result.get('rec_scores', [1.0] * len(texts))
#
#             for i, (text, poly) in enumerate(zip(texts, polys)):
#                 confidence = scores[i] if i < len(scores) else 1.0
#
#                 x_coords = [point[0] for point in poly]
#                 y_coords = [point[1] for point in poly]
#
#                 x_min, x_max = min(x_coords), max(x_coords)
#                 y_min, y_max = min(y_coords), max(y_coords)
#
#                 ocr_data.append({
#                     'text': text.strip(),
#                     'confidence': confidence,
#                     'x_min': x_min,
#                     'x_max': x_max,
#                     'y_min': y_min,
#                     'y_max': y_max,
#                     'x_center': (x_min + x_max) / 2,
#                     'y_center': (y_min + y_max) / 2,
#                     'width': x_max - x_min,
#                     'height': y_max - y_min,
#                     'source': 'paddle'
#                 })
#
#     print(f"✅ PaddleOCR found {len(ocr_data)} text regions")
#
#     # Enhance with TrOCR
#     if use_trocr and ocr_data:
#         print("→ Enhancing with TrOCR...")
#         trocr = TrOCRExtractor()
#         image = cv2.imread(preprocessed_image)
#
#         enhanced_count = 0
#         for item in ocr_data:
#             # Use TrOCR for low-confidence detections or important regions
#             if item['confidence'] < 0.85 or item['y_center'] < 300:  # Header region
#                 bbox = [item['x_min'], item['y_min'], item['x_max'], item['y_max']]
#                 trocr_text = trocr.extract_text_from_region(image, bbox)
#
#                 # If TrOCR gives better result, use it
#                 if trocr_text and len(trocr_text) > len(item['text']) * 0.5:
#                     if item['confidence'] < 0.85:
#                         print(f"   🔄 Enhanced: '{item['text']}' → '{trocr_text}' (conf: {item['confidence']:.2f})")
#                         item['text'] = trocr_text
#                         item['confidence'] = 0.95  # TrOCR result
#                         item['source'] = 'trocr'
#                         enhanced_count += 1
#
#         print(f"✅ TrOCR enhanced {enhanced_count} detections")
#
#     return ocr_data
#
#
# # === EXTRACT HEADER INFORMATION ===
# def extract_header_info(ocr_data):
#     """Robust position-based extraction of Name, Sl. No, and Date."""
#     header_info = {"name": "", "sl_no": "", "date": ""}
#     header_zone = [item for item in ocr_data if item["y_center"] < 300]
#
#     print("\n=== DEBUG: Header Zone Elements (y < 300) ===")
#     for i, item in enumerate(header_zone):
#         source_tag = f"[{item.get('source', 'paddle')}]"
#         print(f"{i:2d}: x={item['x_center']:6.1f} y={item['y_center']:6.1f} {source_tag:8} | '{item['text']}'")
#
#     # NAME EXTRACTION (Left side)
#     name_candidates = []
#     for item in header_zone:
#         x, y = item['x_center'], item['y_center']
#         text = item['text'].strip()
#         text_lower = text.lower()
#
#         if x < 300 and 80 < y < 220:
#             if not re.search(r'[a-zA-Z]', text) or len(text) <= 1:
#                 continue
#
#             clean_text = text.replace('.', '').strip()
#             exclude = ['darpan', 'glass', 'ply', 'concepts', 'email',
#                        'phone', 'contact', 'www', '.com', 'sl', 'no',
#                        'date', 'bill', 'mrp', 'particulars', 'qty',
#                        'rate', 'total', '080', '297']
#
#             is_noise = any(kw in text_lower for kw in exclude)
#
#             if not is_noise and len(clean_text) >= 3:
#                 score = len(clean_text)
#                 if 40 <= x <= 150:
#                     score += 5
#                 if 90 <= y <= 180:
#                     score += 3
#                 if item.get('source') == 'trocr':
#                     score += 2  # Bonus for TrOCR
#
#                 name_candidates.append({
#                     'text': clean_text,
#                     'score': score,
#                     'x': x,
#                     'y': y
#                 })
#
#     if name_candidates:
#         best = max(name_candidates, key=lambda c: c['score'])
#         header_info['name'] = best['text']
#         print(f"\n✅ Selected name: '{best['text']}'")
#
#     # SL. NO EXTRACTION
#     for i, item in enumerate(header_zone):
#         text_lower = item['text'].lower().replace(' ', '').replace('.', '')
#
#         if ('sl' in text_lower or 'si' in text_lower) and 'no' in text_lower:
#             for j in range(i + 1, min(i + 6, len(header_zone))):
#                 next_item = header_zone[j]
#                 next_text = next_item['text'].strip()
#
#                 if re.match(r'^\d{2,6}$', next_text) and next_item['x_center'] > 700:
#                     header_info['sl_no'] = next_text
#                     print(f"✅ Found Sl. No: {next_text}")
#                     break
#             if header_info['sl_no']:
#                 break
#
#     # Fallback for Sl. No
#     if not header_info['sl_no']:
#         for item in header_zone:
#             if item['x_center'] > 800 and item['y_center'] < 150:
#                 text = item['text'].strip()
#                 if re.match(r'^\d{2,6}$', text):
#                     header_info['sl_no'] = text
#                     break
#
#     # DATE EXTRACTION
#     for item in header_zone:
#         x = item['x_center']
#         text = item['text'].strip()
#         text_lower = text.lower()
#
#         if 'date' in text_lower and x > 600:
#             date_match = re.search(r'\.?(\d{1,2})[\|/\.\s]*(\d{1,2})[\|/\.\s]*(\d{2,4})', text)
#
#             if date_match:
#                 day, month, year = date_match.groups()
#                 if len(year) == 2:
#                     year = '20' + year if int(year) < 50 else '19' + year
#
#                 header_info['date'] = f"{day}/{month}/{year}"
#                 print(f"✅ Extracted date: {header_info['date']}")
#                 break
#
#     # Date fallback
#     if not header_info['date']:
#         for item in header_zone:
#             if item['x_center'] > 700 and item['y_center'] < 200:
#                 date_match = re.search(r'\.?(\d{1,2})[\|/\.\s]*(\d{1,2})[\|/\.\s]*(\d{2,4})', item['text'])
#                 if date_match:
#                     day, month, year = date_match.groups()
#                     if len(year) == 2:
#                         year = '20' + year
#                     header_info['date'] = f"{day}/{month}/{year}"
#                     break
#
#     print(f"\n{'=' * 60}")
#     print(f"FINAL EXTRACTED HEADER:")
#     print(f"  Name:   '{header_info['name']}'")
#     print(f"  Sl. No: '{header_info['sl_no']}'")
#     print(f"  Date:   '{header_info['date']}'")
#     print(f"{'=' * 60}\n")
#
#     return header_info
#
#
# # === FIND TABLE HEADER ===
# def find_table_start(data):
#     """Find where the table data starts"""
#     for i, item in enumerate(data):
#         text = item['text'].lower().strip()
#         if 'particulars' in text or ('qty' in text and 'rate' in text):
#             print(f"\nFound table header at element {i}: '{item['text']}'")
#             return i + 5
#     return 15
#
#
# # === GROUP INTO ROWS ===
# def group_into_rows(data, y_threshold=25):
#     """Group OCR elements into rows based on y-coordinate proximity"""
#     if not data:
#         return []
#
#     data_sorted = sorted(data, key=lambda x: x['y_center'])
#     rows = []
#     current_row = [data_sorted[0]]
#     last_y = data_sorted[0]['y_center']
#
#     for item in data_sorted[1:]:
#         if abs(item['y_center'] - last_y) <= y_threshold:
#             current_row.append(item)
#         else:
#             if current_row:
#                 current_row.sort(key=lambda x: x['x_center'])
#                 rows.append(current_row)
#             current_row = [item]
#             last_y = item['y_center']
#
#     if current_row:
#         current_row.sort(key=lambda x: x['x_center'])
#         rows.append(current_row)
#
#     return rows
#
#
# # === SPLIT QTY AND RATE IF COMBINED ===
# def split_qty_rate(text):
#     """Split combined qty and rate strings"""
#     if not text or text.strip() == '':
#         return '', ''
#
#     text = text.strip()
#
#     if '  ' in text:
#         parts = re.split(r'\s{2,}', text)
#         if len(parts) >= 2:
#             return parts[0].strip(), ' '.join(parts[1:]).strip()
#
#     match = re.match(r'^(\d+[a-zA-Z]*)[€$£¥](\d+)$', text)
#     if match:
#         return match.group(1), match.group(2)
#
#     match = re.match(r'^(\d+[a-zA-Z]+)(\d+)$', text)
#     if match:
#         return match.group(1), match.group(2)
#
#     if ' ' in text:
#         parts = text.split()
#         if len(parts) >= 2:
#             return parts[0], ' '.join(parts[1:])
#
#     if re.match(r'^\d+[a-zA-Z]+$', text):
#         return text, ''
#
#     return text, ''
#
#
# # === ASSIGN COLUMNS BASED ON X-POSITION ===
# def assign_to_columns(row_elements):
#     """Assign elements to columns based on x-position"""
#     has_typical_particulars = any(150 <= elem['x_center'] < 500 for elem in row_elements)
#
#     columns = {
#         'mrp': '',
#         'particulars': '',
#         'qty_rate': '',
#         'total': ''
#     }
#
#     items_500_660 = []
#     items_660_850 = []
#
#     for elem in row_elements:
#         x = elem['x_center']
#         text = elem['text'].strip()
#
#         if x < 150:
#             columns['mrp'] = columns['mrp'] + ' ' + text if columns['mrp'] else text
#         elif x < 500:
#             columns['particulars'] = columns['particulars'] + ' ' + text if columns['particulars'] else text
#         elif x < 660:
#             items_500_660.append(text)
#         elif x < 850:
#             items_660_850.append(text)
#         else:
#             columns['total'] = columns['total'] + ' ' + text if columns['total'] else text
#
#     if not has_typical_particulars and items_500_660:
#         columns['particulars'] = items_500_660[0]
#         if len(items_500_660) > 1:
#             columns['qty_rate'] = ' '.join(items_500_660[1:])
#     else:
#         if items_500_660:
#             columns['qty_rate'] = ' '.join(items_500_660)
#
#     if items_660_850:
#         if columns['qty_rate']:
#             columns['qty_rate'] = columns['qty_rate'] + ' ' + ' '.join(items_660_850)
#         else:
#             columns['qty_rate'] = ' '.join(items_660_850)
#
#     columns = {k: v.strip() for k, v in columns.items()}
#     qty, rate = split_qty_rate(columns['qty_rate'])
#
#     return [
#         columns['mrp'],
#         columns['particulars'],
#         qty,
#         rate,
#         columns['total']
#     ]
#
#
# # === MAIN EXECUTION ===
# if __name__ == "__main__":
#     # CONFIG
#     input_image = 'C:/Users/lenovo/Downloads/bill1.jpeg'
#     preprocessed_image = 'C:/Users/lenovo/Downloads/bill1_ocr_ready_combined.jpeg'
#     save_dir = 'C:/Users/lenovo/Desktop/bill1-output'
#     os.makedirs(save_dir, exist_ok=True)
#
#     # STEP 1: PREPROCESS IMAGE
#     print("=== Starting Image Preprocessing ===")
#     preprocess_success = preprocess_bill_image_ocr(input_image, preprocessed_image,
#                                                    denoise_strength=5, apply_otsu=False)
#
#     if not preprocess_success:
#         print("❌ Preprocessing failed! Exiting.")
#         exit(1)
#
#     # STEP 2: RUN ENSEMBLE OCR (PaddleOCR + TrOCR)
#     ocr_data = run_ensemble_ocr(input_image, preprocessed_image, use_trocr=True)
#
#     if not ocr_data:
#         print("❌ No OCR data extracted!")
#         exit(1)
#
#     ocr_data.sort(key=lambda x: (x['y_center'], x['x_center']))
#     print(f"Total OCR elements: {len(ocr_data)}")
#
#     # STEP 3: EXTRACT HEADER INFORMATION
#     print("\n=== Extracting Header Information ===")
#     header_info = extract_header_info(ocr_data)
#
#     # STEP 4: PROCESS TABLE DATA
#     print("\n=== Processing Table Data ===")
#     table_start = find_table_start(ocr_data)
#     table_data = ocr_data[table_start:]
#     print(f"Table data starts at element {table_start}, {len(table_data)} elements remaining")
#
#     table_rows = group_into_rows(table_data)
#     print(f"Grouped into {len(table_rows)} rows")
#
#     # STEP 5: PROCESS ROWS INTO TABLE
#     column_names = ['MRP', 'Particulars', 'Qty', 'Rate', 'Total']
#     processed_rows = []
#
#     print(f"\n=== Processing {len(table_rows)} Rows ===")
#
#     for row_idx, row_elements in enumerate(table_rows):
#         row_text = ' '.join([elem['text'] for elem in row_elements]).lower()
#
#         if any(header in row_text for header in ['particulars', 'qty', 'rate', 'total']) and row_idx < 3:
#             continue
#
#         if any(footer in row_text for footer in ['signature', 'total']) and 'sub' not in row_text:
#             if row_text.count('total') > 0 and row_text.count('sub') == 0:
#                 continue
#
#         if len(row_text.strip()) < 2:
#             continue
#
#         row_data = assign_to_columns(row_elements)
#
#         if row_data[1] or row_data[4]:
#             processed_rows.append(row_data)
#             print(f"Row {len(processed_rows):2d}: Part='{row_data[1][:30]:30}' | Qty='{row_data[2][:10]:10}'")
#
#     # STEP 6: CREATE AND SAVE DATAFRAME
#     if processed_rows:
#         df = pd.DataFrame(processed_rows, columns=column_names)
#         df = df.replace('', np.nan)
#         df_to_save = df[['Particulars', 'Qty']].copy()
#
#         excel_path = os.path.join(save_dir, 'extracted_bill_combined.xlsx')
#         csv_path = os.path.join(save_dir, 'extracted_bill_combined.csv')
#
#         with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
#             header_row = pd.DataFrame([[
#                 f"Name : {header_info['name']}",
#                 '',
#                 f"Sl. No: {header_info['sl_no']}",
#                 '',
#                 f"Date : {header_info['date']}"
#             ]], columns=['A', 'B', 'C', 'D', 'E'])
#
#             header_row.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=0)
#             pd.DataFrame([['', '', '', '', '']], columns=['A', 'B', 'C', 'D', 'E']).to_excel(
#                 writer, sheet_name='Sheet1', index=False, header=False, startrow=1)
#             df_to_save.to_excel(writer, sheet_name='Sheet1', index=False, startrow=2)
#
#         wb = load_workbook(excel_path)
#         ws = wb.active
#
#         for col in range(1, 6):
#             cell = ws.cell(row=1, column=col)
#             cell.font = Font(bold=True)
#             cell.alignment = Alignment(horizontal='left')
#
#         for col in range(1, 3):
#             cell = ws.cell(row=3, column=col)
#             cell.font = Font(bold=True)
#
#         wb.save(excel_path)
#
#         with open(csv_path, 'w', encoding='utf-8') as f:
#             f.write(f"Name: {header_info['name']},,,Sl. No: {header_info['sl_no']},,,Date: {header_info['date']}\n")
#             f.write("\n")
#             df_to_save.to_csv(f, index=False)
#
#         print(f"\n✅ EXTRACTION COMPLETE!")
#         print(f"→ Total rows: {len(df_to_save)}")
#         print(f"→ Excel: {excel_path}")
#         print(f"→ CSV: {csv_path}")
#         print(f"\n=== Header Information ===")
#         print(f"Name: {header_info['name']}")
#         print(f"Sl. No: {header_info['sl_no']}")
#         print(f"Date: {header_info['date']}")
#
#         print(f"\n=== Data Quality Analysis ===")
#         for col in ['Particulars', 'Qty']:
#             non_empty = df_to_save[col].notna().sum()
#             print(f"{col:15}: {non_empty:2d}/{len(df_to_save):2d} ({non_empty / len(df_to_save) * 100:5.1f}%) filled")
#
#         print(f"\n=== Preview ===")
#         print(df_to_save.to_string(index=False))
#
#     else:
#         print("❌
#we cannot install detectron2 as it doesn't supports windows, only supports linux and mac


#craft+paddleocr
# import cv2
# import os
# from paddleocr import PaddleOCR
# import pandas as pd
# import re
# import numpy as np
# from openpyxl import load_workbook
# from openpyxl.styles import Font, Alignment
# from craft_text_detector import Craft
# from PIL import Image
#
# # === IMAGE PREPROCESSING FUNCTION ===
# def preprocess_bill_image_ocr(input_path, output_path, denoise_strength=5, apply_otsu=False):
#     """OCR-friendly preprocessing for bill images."""
#     img = cv2.imread(input_path)
#     if img is None:
#         print(f"❌ Error: Could not read image from {input_path}")
#         return False
#
#     gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
#     denoised = cv2.fastNlMeansDenoising(gray, None, h=denoise_strength, templateWindowSize=7, searchWindowSize=21)
#
#     if apply_otsu:
#         _, processed = cv2.threshold(denoised, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
#     else:
#         processed = denoised
#
#     os.makedirs(os.path.dirname(output_path), exist_ok=True)
#     cv2.imwrite(output_path, processed)
#     print(f"✅ Preprocessed image saved to {output_path}")
#     return True
#
#
# # === CRAFT + PADDLEOCR INTEGRATION ===
# def run_craft_then_paddleocr(preprocessed_image_path, use_craft=True, craft_cuda=False):
#     """
#     Detect text regions with CRAFT (craft-text-detector) and run PaddleOCR on each region.
#     Returns list of dicts with keys: text, confidence, x_min, x_max, y_min, y_max, x_center, y_center, width, height, source
#     """
#     print("\n=== Running CRAFT (text detection) + PaddleOCR (recognition) ===")
#
#     # Initialize PaddleOCR
#     paddle_ocr = PaddleOCR(use_angle_cls=True, lang='en')  # keep this global if you call multiple times
#
#     # Read original preprocessed image
#     image = cv2.imread(preprocessed_image_path)
#     if image is None:
#         print(f"❌ Error: Could not read preprocessed image {preprocessed_image_path}")
#         return []
#
#     h_img, w_img = image.shape[:2]
#
#     ocr_data = []
#
#     if use_craft:
#         craft = Craft(output_dir=None, crop_type="box", cuda=craft_cuda)  # output_dir=None -> no files saved
#         try:
#             craft_result = craft.detect_text(preprocessed_image_path)
#             # craft_result is a dict; common keys: 'boxes', 'polys', 'rotated_boxes', 'cropped_regions'
#             boxes = None
#             if isinstance(craft_result, dict):
#                 # boxes usually Nx4x2
#                 if 'boxes' in craft_result and craft_result['boxes'] is not None:
#                     boxes = craft_result['boxes']
#                 elif 'polys' in craft_result and craft_result['polys'] is not None:
#                     boxes = craft_result['polys']
#                 elif 'rotated_boxes' in craft_result and craft_result['rotated_boxes'] is not None:
#                     boxes = craft_result['rotated_boxes']
#
#             if boxes is None:
#                 print("⚠ CRAFT returned no boxes")
#                 craft.unload_craft()
#                 return []
#
#             # boxes may be numpy array; iterate
#             for b in boxes:
#                 # b is 4 points [[x1,y1],[x2,y2],...]
#                 try:
#                     xs = [int(pt[0]) for pt in b]
#                     ys = [int(pt[1]) for pt in b]
#                 except Exception:
#                     # sometimes box might be flat list; try to reshape
#                     arr = np.array(b).reshape(-1, 2)
#                     xs = [int(pt[0]) for pt in arr]
#                     ys = [int(pt[1]) for pt in arr]
#
#                 x_min = max(0, min(xs))
#                 x_max = min(w_img - 1, max(xs))
#                 y_min = max(0, min(ys))
#                 y_max = min(h_img - 1, max(ys))
#
#                 # safety check for empty
#                 if x_max <= x_min or y_max <= y_min:
#                     continue
#
#                 # Crop region with small padding
#                 pad = 3
#                 x0 = max(0, x_min - pad)
#                 y0 = max(0, y_min - pad)
#                 x1 = min(w_img - 1, x_max + pad)
#                 y1 = min(h_img - 1, y_max + pad)
#
#                 cropped = image[y0:y1, x0:x1]
#
#                 if cropped.size == 0:
#                     continue
#
#                 # PaddleOCR can accept numpy array (BGR)
#                 try:
#                     # paddleocr returns format: list of lines -> each item [[bbox_points], (text, confidence)]
#                     rec_results = paddle_ocr.ocr(cropped, cls=True)
#                 except Exception as e:
#                     # fallback to passing path, but here we have array
#                     rec_results = []
#
#                 # parse paddle result
#                 text = ""
#                 confidence = 0.0
#
#                 # rec_results may be nested differently across versions; handle common structures
#                 if isinstance(rec_results, list) and len(rec_results) > 0:
#                     # sometimes returns [[(box),(text,score)], ...] or [ [ [box], (text, score) ], ... ]
#                     # pick the longest text among detections in the crop
#                     candidates = []
#                     for rr in rec_results:
#                         try:
#                             if isinstance(rr, list) and len(rr) >= 2:
#                                 # rr[1] should be (text, score)
#                                 pred = rr[1]
#                                 if isinstance(pred, tuple) or isinstance(pred, list):
#                                     cand_text = str(pred[0]).strip()
#                                     cand_conf = float(pred[1]) if len(pred) > 1 else 0.0
#                                 elif isinstance(pred, dict) and 'text' in pred:
#                                     cand_text = pred.get('text', '').strip()
#                                     cand_conf = float(pred.get('confidence', 0.0))
#                                 else:
#                                     cand_text = ""
#                                     cand_conf = 0.0
#                                 if cand_text:
#                                     candidates.append((cand_text, cand_conf))
#                         except Exception:
#                             continue
#
#                     if candidates:
#                         # pick best by confidence then length
#                         candidates = sorted(candidates, key=lambda x: (x[1], len(x[0])), reverse=True)
#                         text, confidence = candidates[0]
#                 # fallback if no rec_results: try crop OCR with full image OCR as last resort
#                 if text == "":
#                     # try full image OCR to gather any nearby text (very slow)
#                     try:
#                         full_res = paddle_ocr.ocr(preprocessed_image_path, cls=True)
#                         # no easy mapping; skip fallback heavy step in normal runs
#                     except Exception:
#                         pass
#
#                 # compute centers
#                 x_center = (x0 + x1) / 2
#                 y_center = (y0 + y1) / 2
#
#                 ocr_data.append({
#                     'text': text.strip(),
#                     'confidence': float(confidence) if confidence is not None else 0.0,
#                     'x_min': x0,
#                     'x_max': x1,
#                     'y_min': y0,
#                     'y_max': y1,
#                     'x_center': x_center,
#                     'y_center': y_center,
#                     'width': x1 - x0,
#                     'height': y1 - y0,
#                     'source': 'craft+paddle'
#                 })
#
#         finally:
#             craft.unload_craft()
#     else:
#         # If not using CRAFT, fall back to running PaddleOCR on full preprocessed image and dividing
#         print("→ CRAFT disabled, running PaddleOCR on whole image")
#         paddle_ocr = PaddleOCR(use_angle_cls=True, lang='en')
#         try:
#             full_res = paddle_ocr.ocr(preprocessed_image_path, cls=True)
#             # parse full_res
#             for rr in full_res:
#                 try:
#                     box = rr[0]
#                     pred = rr[1]
#                     xs = [int(pt[0]) for pt in box]
#                     ys = [int(pt[1]) for pt in box]
#                     x_min = min(xs); x_max = max(xs); y_min = min(ys); y_max = max(ys)
#                     text = pred[0] if isinstance(pred, (list, tuple)) else (pred.get('text','') if isinstance(pred, dict) else '')
#                     conf = float(pred[1]) if isinstance(pred, (list,tuple)) and len(pred) > 1 else float(pred.get('confidence', 0.0) if isinstance(pred, dict) else 0.0)
#                     ocr_data.append({
#                         'text': str(text).strip(),
#                         'confidence': conf,
#                         'x_min': x_min,
#                         'x_max': x_max,
#                         'y_min': y_min,
#                         'y_max': y_max,
#                         'x_center': (x_min + x_max) / 2,
#                         'y_center': (y_min + y_max) / 2,
#                         'width': x_max - x_min,
#                         'height': y_max - y_min,
#                         'source': 'paddle'
#                     })
#                 except Exception:
#                     continue
#         except Exception as e:
#             print("PaddleOCR full-image error:", e)
#
#     print(f"✅ CRAFT+PaddleOCR found {len(ocr_data)} text regions")
#     return ocr_data
#
#
# # === EXTRACT HEADER INFORMATION ===
# def extract_header_info(ocr_data):
#     """Robust position-based extraction of Name, Sl. No, and Date."""
#     header_info = {"name": "", "sl_no": "", "date": ""}
#     header_zone = [item for item in ocr_data if item["y_center"] < 300]
#
#     print("\n=== DEBUG: Header Zone Elements (y < 300) ===")
#     for i, item in enumerate(header_zone):
#         source_tag = f"[{item.get('source', 'paddle')}]"
#         print(f"{i:2d}: x={item['x_center']:6.1f} y={item['y_center']:6.1f} {source_tag:12} | '{item['text']}'")
#
#     # NAME EXTRACTION (Left side)
#     name_candidates = []
#     for item in header_zone:
#         x, y = item['x_center'], item['y_center']
#         text = item['text'].strip()
#         text_lower = text.lower()
#
#         if x < 300 and 80 < y < 220:
#             if not re.search(r'[a-zA-Z]', text) or len(text) <= 1:
#                 continue
#
#             clean_text = text.replace('.', '').strip()
#             exclude = ['darpan', 'glass', 'ply', 'concepts', 'email',
#                        'phone', 'contact', 'www', '.com', 'sl', 'no',
#                        'date', 'bill', 'mrp', 'particulars', 'qty',
#                        'rate', 'total', '080', '297']
#
#             is_noise = any(kw in text_lower for kw in exclude)
#
#             if not is_noise and len(clean_text) >= 3:
#                 score = len(clean_text)
#                 if 40 <= x <= 150:
#                     score += 5
#                 if 90 <= y <= 180:
#                     score += 3
#
#                 # no trocr bonus now
#                 name_candidates.append({
#                     'text': clean_text,
#                     'score': score,
#                     'x': x,
#                     'y': y
#                 })
#
#     if name_candidates:
#         best = max(name_candidates, key=lambda c: c['score'])
#         header_info['name'] = best['text']
#         print(f"\n✅ Selected name: '{best['text']}'")
#
#     # SL. NO EXTRACTION
#     for i, item in enumerate(header_zone):
#         text_lower = item['text'].lower().replace(' ', '').replace('.', '')
#
#         if ('sl' in text_lower or 'si' in text_lower) and 'no' in text_lower:
#             for j in range(i + 1, min(i + 6, len(header_zone))):
#                 next_item = header_zone[j]
#                 next_text = next_item['text'].strip()
#
#                 if re.match(r'^\d{2,6}$', next_text) and next_item['x_center'] > 700:
#                     header_info['sl_no'] = next_text
#                     print(f"✅ Found Sl. No: {next_text}")
#                     break
#             if header_info['sl_no']:
#                 break
#
#     # Fallback for Sl. No
#     if not header_info['sl_no']:
#         for item in header_zone:
#             if item['x_center'] > 800 and item['y_center'] < 150:
#                 text = item['text'].strip()
#                 if re.match(r'^\d{2,6}$', text):
#                     header_info['sl_no'] = text
#                     break
#
#     # DATE EXTRACTION
#     for item in header_zone:
#         x = item['x_center']
#         text = item['text'].strip()
#         text_lower = text.lower()
#
#         if 'date' in text_lower and x > 600:
#             date_match = re.search(r'\.?(\d{1,2})[\|/\.\s]*(\d{1,2})[\|/\.\s]*(\d{2,4})', text)
#
#             if date_match:
#                 day, month, year = date_match.groups()
#                 if len(year) == 2:
#                     year = '20' + year if int(year) < 50 else '19' + year
#
#                 header_info['date'] = f"{day}/{month}/{year}"
#                 print(f"✅ Extracted date: {header_info['date']}")
#                 break
#
#     # Date fallback
#     if not header_info['date']:
#         for item in header_zone:
#             if item['x_center'] > 700 and item['y_center'] < 200:
#                 date_match = re.search(r'\.?(\d{1,2})[\|/\.\s]*(\d{1,2})[\|/\.\s]*(\d{2,4})', item['text'])
#                 if date_match:
#                     day, month, year = date_match.groups()
#                     if len(year) == 2:
#                         year = '20' + year
#                     header_info['date'] = f"{day}/{month}/{year}"
#                     break
#
#     print(f"\n{'=' * 60}")
#     print(f"FINAL EXTRACTED HEADER:")
#     print(f"  Name:   '{header_info['name']}'")
#     print(f"  Sl. No: '{header_info['sl_no']}'")
#     print(f"  Date:   '{header_info['date']}'")
#     print(f"{'=' * 60}\n")
#
#     return header_info
#
#
# # === FIND TABLE HEADER ===
# def find_table_start(data):
#     """Find where the table data starts"""
#     for i, item in enumerate(data):
#         text = item['text'].lower().strip()
#         if 'particulars' in text or ('qty' in text and 'rate' in text):
#             print(f"\nFound table header at element {i}: '{item['text']}'")
#             return i + 5
#     return 15
#
#
# # === GROUP INTO ROWS ===
# def group_into_rows(data, y_threshold=25):
#     """Group OCR elements into rows based on y-coordinate proximity"""
#     if not data:
#         return []
#
#     data_sorted = sorted(data, key=lambda x: x['y_center'])
#     rows = []
#     current_row = [data_sorted[0]]
#     last_y = data_sorted[0]['y_center']
#
#     for item in data_sorted[1:]:
#         if abs(item['y_center'] - last_y) <= y_threshold:
#             current_row.append(item)
#         else:
#             if current_row:
#                 current_row.sort(key=lambda x: x['x_center'])
#                 rows.append(current_row)
#             current_row = [item]
#             last_y = item['y_center']
#
#     if current_row:
#         current_row.sort(key=lambda x: x['x_center'])
#         rows.append(current_row)
#
#     return rows
#
#
# # === SPLIT QTY AND RATE IF COMBINED ===
# def split_qty_rate(text):
#     """Split combined qty and rate strings"""
#     if not text or text.strip() == '':
#         return '', ''
#
#     text = text.strip()
#
#     if '  ' in text:
#         parts = re.split(r'\s{2,}', text)
#         if len(parts) >= 2:
#             return parts[0].strip(), ' '.join(parts[1:]).strip()
#
#     match = re.match(r'^(\d+[a-zA-Z]*)[€$£¥](\d+)$', text)
#     if match:
#         return match.group(1), match.group(2)
#
#     match = re.match(r'^(\d+[a-zA-Z]+)(\d+)$', text)
#     if match:
#         return match.group(1), match.group(2)
#
#     if ' ' in text:
#         parts = text.split()
#         if len(parts) >= 2:
#             return parts[0], ' '.join(parts[1:])
#
#     if re.match(r'^\d+[a-zA-Z]+$', text):
#         return text, ''
#
#     return text, ''
#
#
# # === ASSIGN COLUMNS BASED ON X-POSITION ===
# def assign_to_columns(row_elements):
#     """Assign elements to columns based on x-position"""
#     has_typical_particulars = any(150 <= elem['x_center'] < 500 for elem in row_elements)
#
#     columns = {
#         'mrp': '',
#         'particulars': '',
#         'qty_rate': '',
#         'total': ''
#     }
#
#     items_500_660 = []
#     items_660_850 = []
#
#     for elem in row_elements:
#         x = elem['x_center']
#         text = elem['text'].strip()
#
#         if x < 150:
#             columns['mrp'] = columns['mrp'] + ' ' + text if columns['mrp'] else text
#         elif x < 500:
#             columns['particulars'] = columns['particulars'] + ' ' + text if columns['particulars'] else text
#         elif x < 660:
#             items_500_660.append(text)
#         elif x < 850:
#             items_660_850.append(text)
#         else:
#             columns['total'] = columns['total'] + ' ' + text if columns['total'] else text
#
#     if not has_typical_particulars and items_500_660:
#         columns['particulars'] = items_500_660[0]
#         if len(items_500_660) > 1:
#             columns['qty_rate'] = ' '.join(items_500_660[1:])
#     else:
#         if items_500_660:
#             columns['qty_rate'] = ' '.join(items_500_660)
#
#     if items_660_850:
#         if columns['qty_rate']:
#             columns['qty_rate'] = columns['qty_rate'] + ' ' + ' '.join(items_660_850)
#         else:
#             columns['qty_rate'] = ' '.join(items_660_850)
#
#     columns = {k: v.strip() for k, v in columns.items()}
#     qty, rate = split_qty_rate(columns['qty_rate'])
#
#     return [
#         columns['mrp'],
#         columns['particulars'],
#         qty,
#         rate,
#         columns['total']
#     ]
#
#
# # === MAIN EXECUTION ===
# if __name__ == "__main__":
#     # CONFIG
#     input_image = 'C:/Users/lenovo/Downloads/bill1.jpeg'
#     preprocessed_image = 'C:/Users/lenovo/Downloads/bill1_ocr_ready_combined.jpeg'
#     save_dir = 'C:/Users/lenovo/Desktop/bill1-output'
#     os.makedirs(save_dir, exist_ok=True)
#
#     # STEP 1: PREPROCESS IMAGE
#     print("=== Starting Image Preprocessing ===")
#     preprocess_success = preprocess_bill_image_ocr(input_image, preprocessed_image,
#                                                    denoise_strength=5, apply_otsu=False)
#
#     if not preprocess_success:
#         print("❌ Preprocessing failed! Exiting.")
#         exit(1)
#
#     # STEP 2: RUN CRAFT + PADDLEOCR
#     ocr_data = run_craft_then_paddleocr(preprocessed_image, use_craft=True, craft_cuda=False)
#
#     if not ocr_data:
#         print("❌ No OCR data extracted!")
#         exit(1)
#
#     ocr_data.sort(key=lambda x: (x['y_center'], x['x_center']))
#     print(f"Total OCR elements: {len(ocr_data)}")
#
#     # STEP 3: EXTRACT HEADER INFORMATION
#     print("\n=== Extracting Header Information ===")
#     header_info = extract_header_info(ocr_data)
#
#     # STEP 4: PROCESS TABLE DATA
#     print("\n=== Processing Table Data ===")
#     table_start = find_table_start(ocr_data)
#     table_data = ocr_data[table_start:]
#     print(f"Table data starts at element {table_start}, {len(table_data)} elements remaining")
#
#     table_rows = group_into_rows(table_data)
#     print(f"Grouped into {len(table_rows)} rows")
#
#     # STEP 5: PROCESS ROWS INTO TABLE
#     column_names = ['MRP', 'Particulars', 'Qty', 'Rate', 'Total']
#     processed_rows = []
#
#     print(f"\n=== Processing {len(table_rows)} Rows ===")
#
#     for row_idx, row_elements in enumerate(table_rows):
#         row_text = ' '.join([elem['text'] for elem in row_elements]).lower()
#
#         if any(header in row_text for header in ['particulars', 'qty', 'rate', 'total']) and row_idx < 3:
#             continue
#
#         if any(footer in row_text for footer in ['signature', 'total']) and 'sub' not in row_text:
#             if row_text.count('total') > 0 and row_text.count('sub') == 0:
#                 continue
#
#         if len(row_text.strip()) < 2:
#             continue
#
#         row_data = assign_to_columns(row_elements)
#
#         if row_data[1] or row_data[4]:
#             processed_rows.append(row_data)
#             print(f"Row {len(processed_rows):2d}: Part='{row_data[1][:30]:30}' | Qty='{row_data[2][:10]:10}'")
#
#     # STEP 6: CREATE AND SAVE DATAFRAME
#     if processed_rows:
#         df = pd.DataFrame(processed_rows, columns=column_names)
#         df = df.replace('', np.nan)
#         df_to_save = df[['Particulars', 'Qty']].copy()
#
#         excel_path = os.path.join(save_dir, 'extracted_bill_combined.xlsx')
#         csv_path = os.path.join(save_dir, 'extracted_bill_combined.csv')
#
#         with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
#             header_row = pd.DataFrame([[
#                 f"Name : {header_info['name']}",
#                 '',
#                 f"Sl. No: {header_info['sl_no']}",
#                 '',
#                 f"Date : {header_info['date']}"
#             ]], columns=['A', 'B', 'C', 'D', 'E'])
#
#             header_row.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=0)
#             pd.DataFrame([['', '', '', '', '']], columns=['A', 'B', 'C', 'D', 'E']).to_excel(
#                 writer, sheet_name='Sheet1', index=False, header=False, startrow=1)
#             df_to_save.to_excel(writer, sheet_name='Sheet1', index=False, startrow=2)
#
#         wb = load_workbook(excel_path)
#         ws = wb.active
#
#         for col in range(1, 6):
#             cell = ws.cell(row=1, column=col)
#             cell.font = Font(bold=True)
#             cell.alignment = Alignment(horizontal='left')
#
#         for col in range(1, 3):
#             cell = ws.cell(row=3, column=col)
#             cell.font = Font(bold=True)
#
#         wb.save(excel_path)
#
#         with open(csv_path, 'w', encoding='utf-8') as f:
#             f.write(f"Name: {header_info['name']},,,Sl. No: {header_info['sl_no']},,,Date: {header_info['date']}\n")
#             f.write("\n")
#             df_to_save.to_csv(f, index=False)
#
#         print(f"\n✅ EXTRACTION COMPLETE!")
#         print(f"→ Total rows: {len(df_to_save)}")
#         print(f"→ Excel: {excel_path}")
#         print(f"→ CSV: {csv_path}")
#         print(f"\n=== Header Information ===")
#         print(f"Name: {header_info['name']}")
#         print(f"Sl. No: {header_info['sl_no']}")
#         print(f"Date: {header_info['date']}")
#
#         print(f"\n=== Data Quality Analysis ===")
#         for col in ['Particulars', 'Qty']:
#             non_empty = df_to_save[col].notna().sum()
#             print(f"{col:15}: {non_empty:2d}/{len(df_to_save):2d} ({non_empty / len(df_to_save) * 100:5.1f}%) filled")
#
#         print(f"\n=== Preview ===")
#         print(df_to_save.to_string(index=False))
#
#     else:
#         print("❌ No table rows extracted.")
